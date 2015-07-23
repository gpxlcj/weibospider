# -*- coding: utf-8 -*-
from __future__ import absolute_import, division, print_function, unicode_literals

import random
import re

from official.weibo_api import short_to_long
from lib.log import lg_debug, lg_info, lg_warning

#excel读写操作
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

from settings.settings import SAVE_FILE_NAME, SHEET_NAME, MONGO_DB

from pymongo import MongoClient

#存储关注结果
def save_follow_data(get_list, session):
    pass 


#存储搜索结果
def save_search_data(get_list, session, location=0, num=0):
# location: 是否要获取目标地址,1获取，0不获取
    wb = load_workbook(SAVE_FILE_NAME)
    ew = ExcelWriter(wb)
    ws = wb.get_sheet_by_name(SHEET_NAME)
    for i in get_list:
        i = i.encode('utf-8')
        i += u''
        if location == 0:
            coordinate = None
        else:
            coordinate = get_location(i, session)
        temp_replace = re.compile('<[\s\S]+?>')
        i = temp_replace.sub(' ', i)
        i += u'\n'
        i = i.decode('string-escape').encode('gbk')
        i = unicode(i, 'unicode-escape')
        i = i.encode('utf-8')
        if coordinate:
            num += 1
            longitude = coordinate[0]
            latitude = coordinate[1]
            ws.cell(row=num+1, column=2).value = longitude
            ws.cell(row=num+1, column=3).value = latitude
            ws.cell(row=num+1, column=1).value = i        
        elif num == 0:
            num += 1
            ws.cell(row=num+1, column=1).value = i
    ew.save(SAVE_FILE_NAME)
    print (num)
    return num
    

#获取微博中地理位置url
def get_location(get_text, session):
    location_re = re.compile(u'''<a class=[\S\s]+?icon_cd_place[\S\s]+?a>''')
    temp_l = location_re.search(get_text)
    if temp_l:
        origin_url = temp_l.group()
        location_re = re.compile(u'''http:.+? ''')
        url = location_re.search(origin_url).group()[:-2]
        url = url.replace('\\', '')
        print (url+'\n')
        coordinate = short_to_long(url, session, random.randint(0, 5))#顺序为经度、纬度
        lg_info(str(coordinate))
        return coordinate 
    else:
        return None



#mongodb存储数据
def save_data_by_db(get_list):
    client = MongoClient(MONGO_DB['address'], MONGO_DB['port'])
    db = client.weibodata
    db = client.get_database(name=MONGO_DB['db_name'])
    #存储根据北京地理位置获得的微博
    collection = db.get_collection(name=MONGO_DB['collection_name'])
    collection = db.BeijingGeo
    if get_list:
        pass
    else:
        get_list = list()
    for wd in get_list:
        try:
            collection.insert_one(wd).inserted_id
            lg_debug(str(True)+':save success'+str(len(get_list)))
        except Exception:
            lg_debug(str(False)+':mongodb save fail')
            lg_warning(Exception.message)
    return True


#if __name__ == '__main__':
#    save_search_data()


