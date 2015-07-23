# -*- coding:utf8 -*-

import openpyxl as pxl
import shapefile as shp
import pymongo
from settings.settings import MONGO_DB

def xls_convert():
    sp = shp.Writer(shp.POINT)
    sp.field('content')
    sp.field('type', 'C', 40)
    wk = pxl.load_workbook('weibo_iesous.xlsx')
    wb = wk.get_sheet_by_name('iesous')
    l = wb.max_row
    for i in range(2, l+1):
        content = wb.cell(column=1, row=i).value
        longitude = wb.cell(column=2, row=i).value
        latitude = wb.cell(column=3, row=i).value
        try:
            longitude = float(longitude)
            latitude = float(latitude)
            sp.point(longitude, latitude)
            sp.record('xxx', 'Point')
        except:
            pass
    sp.save()

def mongo_convert():
    client = pymongo.MongoClient(MONGO_DB['address'], MONGO_DB['port'])
    db = client.get_database(name=MONGO_DB['db_name'])
    collection = db.get_collection(name=MONGO_DB['collection_name'])
    data = collection.find({})
    sp = shp.Writer(shp.POINT)
    sp.field('content')
    sp.field('type', 'C', 40)
    for i in data:
        try:
            latitude = float(i['geo']['coordinates'][0])
            longitude = float(i['geo']['coordinates'][1])
            sp.point(longitude, latitude)
            sp.record('xxx', 'Point')
        except:
            pass
    sp.save()

if __name__ == '__main__':
    mongo_convert()