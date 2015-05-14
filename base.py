# -*- coding: utf-8 -*-
from __future__ import absolute_import, division, print_function, unicode_literals

import os
import sys

#excel读写操作
from openpyxl import Workbook, load_workbook

#初始化环境
def init_env():
    reload(sys)
    sys.setdefaultencoding('utf8')


#存储抓取的页面信息
def save_catch_page(get_text):
    file_a = open('page_1', 'wb')
    get_text = u'' + get_text
    file_a.write(get_text)
    file_a.close()
    return get_text
    

#初始化excel表格
def init_xls(r_id=0):
    wb = Workbook()
    ws = wb.get_active_sheet()
    if r_id == 0:
        ws.cell(row=1, column=1).value = u'描述'
        ws.cell(row=1, column=2).value = u'经度'
        ws.cell(row=1, column=3).value = u'纬度'
        ws.title = SHEET_NAME
    if r_id == 1:
        ws.cell(row=1, column=1).value = u'关注者'
        ws.cell(row=1, column=2).value = u'关注者id'
        ws.cell(row=1, column=3).value = u'被关注者'
        ws.cell(row=1, column=4).value = u'被关注者id'
    wb.save(SAVE_FILE_NAME)
